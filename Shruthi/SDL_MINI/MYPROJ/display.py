# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *
  
# globally declare wb and sheet variable 
  
# opening the existing excel file 
#wb = load_workbook('C:\\Users\\SHRUTHI BHAT\\Desktop\\Mks.xlsx') 
  
# create the sheet object 
#sheet = wb.active 
  
  '''
def disp(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
  
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Roll-No"
    sheet.cell(row=1, column=3).value = "DBMS"
    sheet.cell(row=1, column=4).value = "TOC"
    sheet.cell(row=1, column=5).value = "CN"
    sheet.cell(row=1, column=6).value = "SEPM"
    sheet.cell(row=1, column=7).value = "ISEE"	'''
  
  '''
# Function to set focus (cursor) 
def focus1(event): 
    # set focus on the course_field box 
    roll_field.focus_set() 
  
  
# Function to set focus 
def focus2(event): 
    # set focus on the sem_field box 
    dbms_field.focus_set() 
  
  
# Function to set focus 
def focus3(event): 
    # set focus on the form_no_field box 
    toc_field.focus_set() 
  
  
# Function to set focus 
def focus4(event): 
    # set focus on the contact_no_field box 
    cn_field.focus_set() 
  
  
# Function to set focus 
def focus5(event): 
    # set focus on the email_id_field box 
    sepm_field.focus_set() 
  
  
# Function to set focus 
def focus6(event): 
    # set focus on the address_field box 
    isee_field.focus_set() 
  
  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box 
    name_field.delete(0, END) 
    roll_field.delete(0, END) 
    dbms_field.delete(0, END) 
    toc_field.delete(0, END) 
    cn_field.delete(0, END) 
    sepm_field.delete(0, END) 
    isee_field.delete(0, END) '''
  
 ''' 
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
      
    # if user not fill any entry 
    # then print "empty input" 
    if (name_field.get() == "" and
        roll_field.get() == "" and
        dbms_field.get() == "" and
        toc_field.get() == "" and
        cn_field.get() == "" and
        sepm_field.get() == "" and
        isee_field.get() == ""): 
              
        print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = roll_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = dbms_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = toc_field.get() 
        sheet.cell(row=current_row + 1, column=5).value = cn_field.get() 
        sheet.cell(row=current_row + 1, column=6).value = sepm_field.get() 
        sheet.cell(row=current_row + 1, column=7).value = isee_field.get() 
  
        # save the file 
        wb.save('C:\\Users\\SHRUTHI BHAT\\Desktop\\Mks.xlsx') 
  
        # set focus on the name_field box 
        name_field.focus_set() 
  
        # call the clear() function 
        clear() 
  '''
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='Red') 
  
    # set the title of GUI window 
    root.title("STUDENT ACADEMIC RECORD") 
  
    # set the configuration of GUI window 
    root.geometry("500x300") 
  
   # excel() 
  
    # create a Form Button 
    AButton = Button(root, text="Overall analysis", fg="Black", bg="White")
    BButton = Button(root, text="HighestMarks" , fg="Black", bg="White")
    CButton = Button(root, text="Lowest Marks", fg="Black", bg="White")

  '''
    # create a Name label 
    name = Label(root, text="Name", bg="Pink") 
  
    # create a Course label 
    roll = Label(root, text="Roll-No", bg="Pink") 
  
    # create a Semester label 
    dbms = Label(root, text="Marks-DBMS", bg="Pink") 
  
    # create a Form No. lable 
    toc = Label(root, text="Marks-TOC", bg="Pink") 
  
    # create a Contact No. label 
    cn = Label(root, text="Marks-CN", bg="Pink") 
  
    # create a Email id label 
    sepm = Label(root, text="Marks-SEPM", bg="Pink") 
  
    # create a address label 
    isee = Label(root, text="Marks-ISEE", bg="Pink") 
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=1) 
    name.grid(row=1, column=0) 
    roll.grid(row=2, column=0) 
    dbms.grid(row=3, column=0) 
    toc.grid(row=4, column=0) 
    cn.grid(row=5, column=0) 
    sepm.grid(row=6, column=0) 
    isee.grid(row=7, column=0) 
  
    # create a text entry box 
    # for typing the information 
    name_field = Entry(root) 
    roll_field = Entry(root) 
    dbms_field = Entry(root) 
    toc_field = Entry(root) 
    cn_field = Entry(root) 
    sepm_field = Entry(root) 
    isee_field = Entry(root) 
  
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
    name_field.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    roll_field.bind("<Return>", focus2) 
  
    # whenever the enter key is pressed 
    # then call the focus3 function 
    dbms_field.bind("<Return>", focus3) 
  
    # whenever the enter key is pressed 
    # then call the focus4 function 
    toc_field.bind("<Return>", focus4) 
  
    # whenever the enter key is pressed 
    # then call the focus5 function 
    cn_field.bind("<Return>", focus5) 
  
    # whenever the enter key is pressed 
    # then call the focus6 function 
    sepm_field.bind("<Return>", focus6) 
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    name_field.grid(row=1, column=1, ipadx="100") 
    roll_field.grid(row=2, column=1, ipadx="100") 
    dbms_field.grid(row=3, column=1, ipadx="100") 
    toc_field.grid(row=4, column=1, ipadx="100") 
    cn_field.grid(row=5, column=1, ipadx="100") 
    sepm_field.grid(row=6, column=1, ipadx="100") 
    isee_field.grid(row=7, column=1, ipadx="100") 
  
    # call excel function 
    disp() 
  '''
    # create a Submit Button and place into the root window 
    #submit = Button(root, text="Submit", fg="Black", 
    #                        bg="white", command=insert) 
   # submit.grid(row=8, column=1) 
  
    # start the GUI 
    root.mainloop() 